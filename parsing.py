import os
import re
import requests
from bs4 import BeautifulSoup
from openpyxl.reader.excel import load_workbook

from DataClasses import ScheduleXls, Location, Institute, Schedule, DaysOfWeek, PairNumber, ParityOfWeek, Pair


def getSchelduleXml(qual, institute, location, course):
    URL = "https://www.mirea.ru/schedule/"
    r = requests.get(URL)
    soup = BeautifulSoup(r.text, "html.parser")
    scheduleMirea = soup.find('ul', id='tab-content').find_all('li', )
    schedule = []

    qualification = ["бакалавриат/специалитет", "магистратура", "аспирантура", "", ""]
    a = 0
    for qualif in scheduleMirea:
        if a == 3:
            break
        a += 1
        schedule.append(ScheduleXls(qualification[len(schedule)], []))
        for inst in qualif.find().find_all(recursive=False):
            if (not inst.find(class_="uk-text-bold")) or inst.find(
                    class_="uk-text-bold").text == "Филиал в городе Ставрополе":
                continue
            schedule[-1].institutes.append(Institute(inst.find(class_="uk-text-bold").text, []))
            for course in inst.find(class_=re.compile("uk-width-1-1 uk-grid-small")).find_all("div"):
                if course['class'][0] != "uk-width-expand@m" and course['class'][0] != "uk-width-1-2":
                    continue
                if course['class'][0] == "uk-width-expand@m":
                    schedule[-1].institutes[-1].location.append(Location(course.text, []))
                    continue
                schedule[-1].institutes[-1].location[-1].courses.append(
                    course.find(class_='uk-link-toggle').get("href"))
    return schedule[qual].institutes[institute].location[location].courses[course]


def scheduleOfGroups(scheduleXML):
    req = requests.get(scheduleXML)
    file = open("schedule.xlsx", "wb")
    file.write(req.content)
    wb = load_workbook('schedule.xlsx')
    daysOfWeek = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]
    scheduleOfTheGroups = []
    for ws in wb:
        for i in range(0, ws.max_column // 5):
            if (i + 1) % 3 == 0:
                continue
            scheduleOfTheGroups.append(Schedule('', []))
            a = 1
            day = 0
            pairNumber = 1
            for row in ws.iter_rows(min_row=2, max_row=87, min_col=6 + 5 * i, max_col=9 + 5 * i, values_only=True):
                if a == 1:
                    scheduleOfTheGroups[-1].group = row[0]
                    a += 1
                    continue
                if a == 2:
                    a += 1
                    continue
                if (a - 3) % 14 == 0:
                    pairNumber = 1
                    scheduleOfTheGroups[-1].dayOfWeek.append(DaysOfWeek(daysOfWeek[day], []))
                    day += 1
                if (a - 3) % 2 == 0:
                    scheduleOfTheGroups[-1].dayOfWeek[-1].pairNumber.append(PairNumber(pairNumber, []))
                    scheduleOfTheGroups[-1].dayOfWeek[-1].pairNumber[-1].parityOfWeek.append(
                        ParityOfWeek(True, Pair(row[0], row[1], row[2], row[3])))
                    pairNumber += 1
                else:
                    scheduleOfTheGroups[-1].dayOfWeek[-1].pairNumber[-1].parityOfWeek.append(
                        ParityOfWeek(True, Pair(row[0], row[1], row[2], row[3])))
                a += 1
    return scheduleOfTheGroups
