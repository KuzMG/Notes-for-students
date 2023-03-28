
import re

import requests
from bs4 import BeautifulSoup
from openpyxl.reader.excel import load_workbook

from database.Table_db import Schedule, DaysOfWeek, PairNumber, ParityOfWeek, Pair, engine
from dto.DataClasses import ScheduleXls, Location, Institute
import docx


def get_schedule_xls():
    URL = "https://www.mirea.ru/schedule/"
    r = requests.get(URL)
    soup = BeautifulSoup(r.text, "html.parser")
    scheduleMirea = soup.find('ul', id='tab-content').find_all('li', )
    schedule = []

    qualification = ["Бакалавриат/Специалитет", "Магистратура"]
    a = 0
    for qualif in scheduleMirea:
        if a == 2:
            break
        a += 1
        schedule.append(ScheduleXls(qualification[len(schedule)], []))
        for inst in qualif.find().find_all(recursive=False):
            if (not inst.find(class_="uk-text-bold")) or inst.find(
                    class_="uk-text-bold").text == "Филиал в городе Ставрополе":
                continue
            schedule[-1].institutes.append(Institute(inst.find(class_="uk-text-bold").text.strip(), []))
            for course in inst.find(class_=re.compile("uk-width-1-1 uk-grid-small")).find_all("div"):
                if course['class'][0] != "uk-width-expand@m" and course['class'][0] != "uk-width-1-2":
                    continue
                if course['class'][0] == "uk-width-expand@m":
                    schedule[-1].institutes[-1].location.append(Location(course.text.strip(), []))
                    continue
                schedule[-1].institutes[-1].location[-1].courses.append(
                    course.find(class_='uk-link-toggle').get("href"))
    return schedule


def schedule_of_groups_xls(schedule_xls):
    req = requests.get(schedule_xls)
    file = open("res/schedule/schedule.xlsx", "wb")
    file.write(req.content)
    wb = load_workbook('res/schedule/schedule.xlsx')
    daysOfWeek = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]
    scheduleOfTheGroups = []
    for ws in wb:
        for i in range(0, ws.max_column // 5):
            if (i + 1) % 3 == 0:
                continue
            scheduleOfTheGroups.append(Schedule(group='', days_of_week=[]))
            a = 1
            day = 0
            pairNumber = 1
            for row in ws.iter_rows(min_row=2, max_row=87, min_col=6 + 5 * i, max_col=9 + 5 * i, values_only=True):
                if a == 1:
                    scheduleOfTheGroups[-1].group = row[0]
                    a += 1
                    continue
                if a == 2:
                    a += 1
                    continue
                if (a - 3) % 14 == 0:
                    pairNumber = 1
                    scheduleOfTheGroups[-1].days_of_week.append(DaysOfWeek(day_of_week=daysOfWeek[day], pair_number=[]))
                    day += 1
                discipline = row[0]
                if discipline is not None and discipline.find("н.") != -1:
                    discipline = discipline[discipline.find("н.") + 3:]
                if (a - 3) % 2 == 0:
                    scheduleOfTheGroups[-1].days_of_week[-1].pair_number.append(
                        PairNumber(pair_number=pairNumber, parity=[]))
                    scheduleOfTheGroups[-1].days_of_week[-1].pair_number[-1].parity.append(
                        ParityOfWeek(parity=True,
                                     pair=[Pair(discipline=discipline, occupation=row[1], name_of_the_teacher=row[2],
                                                number_of_cabinet=row[3])]))
                    pairNumber += 1
                else:
                    scheduleOfTheGroups[-1].days_of_week[-1].pair_number[-1].parity.append(
                        ParityOfWeek(parity=False,
                                     pair=[Pair(discipline=discipline, occupation=row[1], name_of_the_teacher=row[2],
                                                number_of_cabinet=row[3])]))
                a += 1
    return scheduleOfTheGroups


def get_schedule_bgu_pdf():
    URL = "https://brgu.ru/sveden/education/raspisanie-zanyatiy/fak-istori-i-mezdynarodnix-otnow/"
    r = requests.get(URL)
    soup = BeautifulSoup(r.text, "html.parser")
    schedule = []
    for course in soup.find('div', class_='col-sm-9 gray').find_all('p'):
        schedule.append("https://brgu.ru" + course.find('a').get('href'))
    return schedule


date_bgu = ['09.00-10.35', '10.45-12.20', '13.00-14.35', '14.45-16.20']


def schedule_of_groups_docx(path):
    doc = docx.Document(path)
    all_tables = doc.tables
    schedule = []
    pattern = re.compile('w:fill=\"(\S*)\"')
    for table in all_tables:
        schedule.append(Schedule(group=None, days_of_week=[]))
        day = 0
        for i, row in enumerate(table.rows):
            if i == 0:
                continue
            if i == 1:
                direct = row.cells[1].text.split("\n")[1]
                schedule[-1].group = direct.strip()
                continue
            if pattern.search(row.cells[-1]._tc.xml).group(1) != "FFFFFF":
                day_of_week = row.cells[-1].text.strip()
                if day < 5:
                    schedule[-1].days_of_week.append(DaysOfWeek(day_of_week=day_of_week, pair_number=[]))
                    last_date = '-'
                day += 1
                continue
            for j, cell in enumerate(row.cells):
                if j == 0:
                    date = cell.text.replace(' ', '')
                    try:
                        pair_number = date_bgu.index(date) + 1
                        date = cell.text
                    except:
                        last_date = schedule[-1].days_of_week[-1].pair_number[-1].date
                        index = date_bgu.index(last_date)
                        pair_number = index + 2
                        date = date_bgu[index + 1]
                    if last_date != date:
                        schedule[-1].days_of_week[-1].pair_number.append(
                            PairNumber(pair_number=pair_number, date=date, parity=[ParityOfWeek(parity=True, pair=[])]))
                    else:
                        schedule[-1].days_of_week[-1].pair_number.append(
                            PairNumber(pair_number=pair_number, date=date,
                                       parity=[ParityOfWeek(parity=False, pair=[])]))
                    last_date = date
                    continue
                text = cell.text.strip().replace('\n', '')
                if j == 1:
                    first_cell = text
                if j == 2:
                    medium_cell = text
                if j == 3:
                    last_cell = text
                len_cell = j
            if len_cell == 3:
                if first_cell != '' and medium_cell != '' and last_cell != '':
                    pair = replace_string(first_cell)
                    schedule[-1].days_of_week[-1].pair_number[-1].parity[-1].pair.append(
                        Pair(discipline=pair[0], occupation=pair[1], name_of_the_teacher=pair[2], number_of_cabinet=''))
                elif first_cell != '':
                    pair = replace_string(first_cell)
                    schedule[-1].days_of_week[-1].pair_number[-1].parity[-1].pair.append(
                        Pair(discipline=pair[0], occupation=pair[1], name_of_the_teacher=pair[2], number_of_cabinet=''))
                elif last_cell != '':
                    pair = replace_string(last_cell)
                    schedule[-1].days_of_week[-1].pair_number[-1].parity[-1].pair.append(
                        Pair(discipline=pair[0], occupation=pair[1], name_of_the_teacher=pair[2], number_of_cabinet=''))
                elif first_cell == '' and medium_cell == '' and last_cell == '':
                    schedule[-1].days_of_week[-1].pair_number[-1].parity[-1].pair.append(
                        Pair(discipline='', occupation='', name_of_the_teacher='', number_of_cabinet=''))
            if len_cell == 1:
                pair = replace_string(first_cell)
                schedule[-1].days_of_week[-1].pair_number[-1].parity[-1].pair.append(
                    Pair(discipline=pair[0], occupation=pair[1], name_of_the_teacher=pair[2], number_of_cabinet=''))
    return schedule


def replace_string(text):
    # text = text.replace("\n", " ")
    # text = text.strip()
    occupation = ''
    name_of_the_teacher = ''
    if text.find('лекция') != -1:
        occupation = text[text.find('лекция'):text.find('лекция') + 6]
        text = text[:text.find('лекция')]
    elif text.find('семинар') != -1:
        occupation = text[text.find('семинар'):text.find('семинар') + 7]
        text = text[:text.find('семинар')]
    elif text.find('П/р') != -1:
        occupation = text[text.find('П/р'):text.find('П/р') + 3]
        text = text[:text.find('П/р')]
    elif text.find('л/р') != -1:
        occupation = text[text.find('л/р'):text.find('л/р') + 3]
        text = text[:text.find('л/р')]
    if text.find('Учебная практика') != -1:
        discipline = text
        occupation = ''
        name_of_the_teacher = ''
    else:
        try:
            name_of_the_teacher = text[text.rindex('(') + 1:text.rindex(')')]
            discipline = text[:text.rindex('(')]
        except:
            discipline = text
    return discipline, occupation, name_of_the_teacher